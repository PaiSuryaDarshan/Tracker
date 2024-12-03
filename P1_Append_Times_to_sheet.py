import openpyxl
from datetime import datetime
from copy import deepcopy
import P0_Retrieve_Active_Sheet

def add_times():
    # Get the active sheet
    returned_list = P0_Retrieve_Active_Sheet.get_active_sheet()  

    filename, wb, sheet, Index_of_Sheet_of_interest = returned_list

    # Grab all tables on the sheet
    tables = []

    for table in sheet.tables.values():
        tables.append(table)

    # Grab loc and style of table of interest (TOI)
    toi = tables[0]
    style = deepcopy(toi.tableStyleInfo)

    # BUG FIX 3: Reset 'max_rows' from "max rows on sheet" to "max rows on sheet with data" 
    Total_Rows = sheet.max_row
    Index_for_cleaning_up = int(toi.ref.split(':')[1].strip('E'))
    sheet.delete_rows(Index_for_cleaning_up,10000)
    cleaned_up_rows = sheet.max_row

    # Create a list to store the values 
    SignIn_Times = [] 
    
    # # Iterate through columns 
    # for column in sheet.iter_cols(): 
    #     # Get the value of the first cell in the 
    #     # column (the cell with the column name) 
    #     column_name = column[0].value 
    #     # Check if the column is the "Name" column 
    #     if column_name == "Sign-in": 
    #         # Iterate over the cells in the column 
    #         for cell in column: 
    #             # Add the value of the cell to the list 
    #             SignIn_Times.append(cell.value) 

    New_additions = 0                                                           #^ File Edit

    print("Accessing Sheet", sheet)

    def Sign_in():
        date = datetime.now().strftime('%d/%m/%y')
        time = datetime.now().strftime('%H:%M')

        print("\n Signing in", date, time)
        
        sheet.append([date, "", time, "", ""])

        toi.ref = f"A1:E{Index_for_cleaning_up + 1}"                                #^ File Edit

        # BUG FIX 1: Integers in table number 10 to cause problems when processing in Microsoft Excel, therefore, table name was adapted to avoid Integers.
        toi.name = f"Table{Index_of_Sheet_of_interest}"                                                          #^ File Edit

        # BUG FIX 2: Integers in table number 10 to cause problems when processing in Microsoft Excel, therefore, table name was adapted to avoid Integers.
        toi.tableStyleInfo = style                                                  #^ File Edit

        wb.save(filename)

    def Sign_out():
        date = datetime.now().strftime('%d/%m/%y')
        time = datetime.now().strftime('%H:%M')

        print("\n Signing out", date, time)
        
        Last_sign_in_cell = "C" + str((int(toi.ref[-1]))-1)
        Last_empty_sign_out_cell = "D" + str((int(toi.ref[-1]))-1)
        Last_empty_time_spent_cell = "E" + str((int(toi.ref[-1]))-1)

        if sheet[Last_empty_sign_out_cell].value == None:
            sheet[Last_empty_sign_out_cell].value = time                               #^ File Edit
            
            # Calculate time spent on task
            Sign_in_time_str = sheet[Last_sign_in_cell].value
            time_format = '%H:%M'
            Sign_in_time = datetime.strptime(Sign_in_time_str, time_format)
            Sign_out_time = datetime.strptime(time, time_format)
            sheet[Last_empty_time_spent_cell].value = Sign_out_time - Sign_in_time
        else:
            print("already signed out")

        # BUG FIX 1: Integers in table number 10 to cause problems when processing in Microsoft Excel, therefore, table name was adapted to avoid Integers.
        toi.name = f"Table{Index_of_Sheet_of_interest}"                                                             #^ File Edit

        # BUG FIX 2: Integers in table number 10 to cause problems when processing in Microsoft Excel, therefore, table name was adapted to avoid Integers.
        toi.tableStyleInfo = style                                                  #^ File Edit

        wb.save(filename)


    append_time = input("\n 1 ----> Sign In \n\n 2 ----> Sign Out")

    if int(append_time) == 1:
        Sign_in()
    elif int(append_time) == 2:
        Sign_out()

    return